from sre_parse import FLAGS
from tkinter import FIRST
from tokenize import group
import requests
import openpyxl
import vk_api
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.utils import get_random_id
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from vk_api import VkUpload
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from translate import Translator
import matplotlib.pyplot as plt
import PIL.Image as Image
import numpy as np


DAYS = {"понедельник": 0, "вторник": 1, "среда": 2, "четверг": 3,
        "пятница": 4, "суббота": 5, "воскресенье": 6}
FIRST_DAY = datetime(2022, 2, 7)
WEATHER_KEY = '0f9e28e70c3287577404f914752dc92d'
URL = "https://www.mirea.ru/schedule/"

def normalize_date(a):
    if a < 10: a = "0" + str(a)
    a = str(a)
    return a
    
def get_day_of_week(day):
    if day.lower() == "сегодня":
        dow = int(datetime.today().weekday())
    elif day.lower() == "завтра":
        dow = (int(datetime.today().weekday()) + 1) % 7
    else:
        dow = DAYS[day.lower()]
    return dow

def get_evenness(day = datetime.today()):
    if day == 'сегодня':
        day = datetime.today()
    elif day == 'завтра':
        day = datetime.today() + timedelta(1)
    week = ((int((day - FIRST_DAY).days)) // 7) % 7 + 1
    if week == 0:
        return 1
    return 0

def get_schedule(group, day, evenness):
    schedule = []
    if day == 6:
        return ['выходной']
    course = get_course(group)
    for i in range(1, NUM_COLS[course]):
        if SHEET[course].cell(2, i).value == group:
            for s in range(4 + 12 * day + evenness, 4 + 12 * (day + 1) + evenness, 2):
                subj = SHEET[course].cell(s, i).value
                if SHEET[course].cell(s, i+2).value:
                    subj += ", " + SHEET[course].cell(s, i+2).value 
                if SHEET[course].cell(s, i+3).value:
                    subj += ", " + SHEET[course].cell(s, i+3).value
                schedule.append(subj) if subj else schedule.append("--")
    return schedule

def get_week_schedule(number_of_group, day):
    DAYS = {0: "Понедельник", 1: "Вторник", 2: "Среда", 3: "Четверг", 4: "Пятница", 5: "Суббота"}
    even = get_evenness(day)
    s = ""
    for i in range(6):
        s += DAYS[i] + ":" + "\n"
        schedule = get_schedule(number_of_group, i, even)
        for j in range(len(schedule)):
            s += str(j + 1) + ") "
            if (schedule[j] == None): s += "----" + "\n"
            else: s += str(schedule[j]) + "\n"
    return s

def get_formatted_schedule(group, day):
    if day.lower() == "эта неделя":
        return get_week_schedule(group, datetime.today())
    if (day.lower() == "следующая неделя"):
        return get_week_schedule(group, datetime.today() + timedelta(7))
    elif day.lower() in ("сегодня", "завтра"):    
        ev = get_evenness(day)
        day = get_day_of_week(day)
        s = get_schedule(group, day, ev)
        out = ""
        print(s)
        for i in range(6):
            out += str(i+1) + ') ' + s[i] + '\n'
        print(out)
        return out
    else:
        s = ''
        day = get_day_of_week(day)
        s += "Расписание для нечетной недели:\n"
        schedule = get_schedule(group, day, 0)
        for i in range(len(schedule)):
            s += str(i+1) + ") " + schedule[i] + '\n'
        s += "_____________________________\n"
        s += "Расписание для четной недели:\n"
        schedule = get_schedule(group, day, 1)
        for i in range(len(schedule)):
            s += str(i+1) + ") " + schedule[i] + '\n'
        return s

def get_teacher_schedule(name):
    schedule = [[['--', '--', '--', '--', '--', '--'] for _ in range(6)] for i in range(2)]
    full_name = ""
    print(name.lower())
    for course in range(3):
        for i in range(4, NUM_ROWS[course]):
            for j in range(8, NUM_COLS[course]):
                try:
                    if name.lower() in SHEET[course].cell(i, j).value.lower():
                        if len(full_name) == 0: full_name = SHEET[course].cell(i, j).value
                        schedule[(i+1)%2][(i - 4) // 12][(i - 2) // 2 % 6 - 1] = ", ".join([(SHEET[course].cell(i, j-2).value), (SHEET[course].cell(i, j-1).value), (SHEET[course].cell(2, j-2).value), (SHEET[course].cell(i, j+1).value)])
                except AttributeError:
                    pass
    return [schedule, full_name]

def get_formatted_teacher(name, day):
    DAYS = {0: "Понедельник", 1: "Вторник", 2: "Среда", 3: "Четверг", 4: "Пятница", 5: "Суббота"}
    if day == "на сегодня":
        d = datetime.today()
        ev = get_evenness(d)
        dof = get_day_of_week("сегодня")
        s = get_teacher_schedule(name)[0][ev][dof]
        out = ''
        for j in range(6):
            out += str(j + 1) + ") "
            out += s[j] + '\n'
        return out
    elif day == "на завтра":
        d = datetime.today() + timedelta(1)
        ev = get_evenness(d)
        dof = get_day_of_week("завтра") 
        s = get_teacher_schedule(name)[0][ev][dof]
        out = ''
        for j in range(6):
            out += str(j + 1) + ") "
            out += s[j] + '\n'
        return out
    elif day == "на эту неделю":
        d = datetime.today()
        ev = get_evenness(d)
        s = get_teacher_schedule(name)[0][ev]
        out = ''
        for i in range(6):
            out += DAYS[i] + ":" + "\n"
            for j in range(6):
                out += str(j + 1) + ") "
                out += s[0][i] + '\n'
        return out
    elif day == "на следующую неделю":
        d = (datetime.today() + timedelta(7))
        ev = get_evenness(d)
        s = get_teacher_schedule(name)[0][ev]
        out = ''
        for i in range(6):
            out += DAYS[i] + ":" + "\n"
            for j in range(6):
                out += str(j + 1) + ") "
                out += s[0][i] + '\n'
        return out

def get_links():
    page = requests.get(URL)
    links = []
    allLinks = []
    soup = BeautifulSoup(page.text, "html.parser")
    allLinks = soup.findAll('a', class_="uk-link-toggle")
    for i in range(len(allLinks)):
        allLinks[i] = allLinks[i].get('href')
    for link in allLinks:
        if 'ИИТ' in link and 'курс_21-22' in link and 'зач' not in link:
            links.append(link)
    return links

def get_course(group):
    return 21 - int(group[-2:])

def call_keyboard_first(keyboard, vk, event):
    keyboard.add_button('Получить расписание', color=VkKeyboardColor.NEGATIVE)
    keyboard.add_line()
    keyboard.add_button('Получить погоду', color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button('Получить статистику по коронавирусу', color=VkKeyboardColor.POSITIVE)
    vk.messages.send(user_id = event.user_id, random_id = get_random_id(), keyboard=keyboard.get_keyboard(), message='Что вы хотите узнать?')

def first_message(vk, event):
    print('New from {}, text = {}'.format(event.user_id, event.text))
    vk.messages.send(
    user_id = event.user_id,
    random_id = get_random_id(),
    message = 'Привет, ' + \
    vk.users.get(user_id = event.user_id)[0]['first_name']
    )
    vk.messages.send(
    user_id = event.user_id,
    random_id = get_random_id(),
    message = 'Напиши БОТ, если хочешь посмотреть функции'
    )

def call_keyboard_schedule(keyboard, vk, event):
    keyboard.add_button('сегодня', color=VkKeyboardColor.POSITIVE)
    keyboard.add_button('завтра', color=VkKeyboardColor.NEGATIVE)
    keyboard.add_line()
    keyboard.add_button('эта неделя', color=VkKeyboardColor.PRIMARY)
    keyboard.add_button('следующая неделя', color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button('какая неделя?', color=VkKeyboardColor.SECONDARY)
    keyboard.add_button('какая группа?', color=VkKeyboardColor.SECONDARY)
    vk.messages.send(user_id = event.user_id, random_id = get_random_id(), keyboard=keyboard.get_keyboard(), message='Выберете варианты')

def call_keyboard_weather(keyboard, vk, event):
    keyboard.add_button('сейчас', color=VkKeyboardColor.POSITIVE)
    keyboard.add_button('сегодня', color=VkKeyboardColor.PRIMARY)
    keyboard.add_button('завтра', color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button('на 5 дней', color=VkKeyboardColor.NEGATIVE)
    vk.messages.send(user_id = event.user_id, random_id = get_random_id(), keyboard=keyboard.get_keyboard(), message='Выберете варианты')

def call_keyboard_teacher(keyboard, vk, event):
    keyboard.add_button('на сегодня', color=VkKeyboardColor.POSITIVE)
    keyboard.add_button('на завтра', color=VkKeyboardColor.NEGATIVE)
    keyboard.add_line()
    keyboard.add_button('на эту неделю', color=VkKeyboardColor.PRIMARY)
    keyboard.add_button('на следующую неделю', color=VkKeyboardColor.PRIMARY)
    vk.messages.send(user_id = event.user_id, random_id = get_random_id(), keyboard=keyboard.get_keyboard(), message='Выберите вариант: ')

def weather_in_time(response, vk_session):
    image = requests.get("http://openweathermap.org/img/wn/{}@2x.png".format(response['weather'][0]['icon']), stream=True)
    information = ""
    result = response["weather"][0]["description"]
    information += result + ", " + "температура: " + str(round(response["main"]["temp_min"])) + " - " + str(round(response["main"]["temp_max"])) + " °C" + "\n"
    information += "Давление: " + str(response["main"]["pressure"]) + " мм.рт.ст., " + "влажность: " + str(response["main"]["humidity"]) + "%" + "\n"
    speed = response["wind"]["speed"]
    deg = response["wind"]["deg"]
    information += "Ветер: " + bofort_scale(speed) + ", " + str(speed) + " м/с, " + rumb(deg)

    return [image, information]

def get_weather_today(day, vk_session, vk, event):
    response = requests.get("http://api.openweathermap.org/data/2.5/forecast?q=moscow&appid=1e7d1c94703c5b863a60ea656e79de92&lang=ru&units=metric")
    info = response.json()
    ms = "Погода в Москве " + day
    data = datetime.today().date()
    information = ""
    if (day == "завтра"): data = (datetime.today() + timedelta(1)).date()
    vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
    counter = 0
    for i in range(len(info["list"])):
        if str(data) in info["list"][i]["dt_txt"]:
            if "6:00:00" in info["list"][i]["dt_txt"]:
                counter += 1
                response = info["list"][i]
                collect = weather_in_time(response, vk_session)
                information += "УТРО:\n"
                attachments = collect[0]
                information += collect[1] + "\n"
                with open("c:/Users/Huawei/Desktop/file1.png", "wb") as f:
                    f.write(attachments.content)
            if "12:00:00" in info["list"][i]["dt_txt"]:
                counter += 1
                response = info["list"][i]
                collect = weather_in_time(response, vk_session)
                attachments = collect[0]
                information += "ДЕНЬ:\n"
                information += collect[1] + "\n"
                with open("c:/Users/Huawei/Desktop/file2.png", "wb") as f:
                    f.write(attachments.content)
            if "18:00:00" in info["list"][i]["dt_txt"]:
                counter += 1
                response = info["list"][i]
                collect = weather_in_time(response, vk_session)
                attachments = collect[0]
                information += "ВЕЧЕР:\n"
                information += collect[1] + "\n"
                with open("c:/Users/Huawei/Desktop/file3.png", "wb") as f:
                    f.write(attachments.content)
            if "21:00:00" in info["list"][i]["dt_txt"]:
                counter += 1
                response = info["list"][i]
                collect = weather_in_time(response, vk_session)
                attachments = collect[0]
                information += "НОЧЬ:\n"
                information += collect[1] + "\n"
                with open("c:/Users/Huawei/Desktop/file4.png", "wb") as f:
                    f.write(attachments.content)
                img = Image.new('RGBA', (counter * 100, 100))
                img4 = Image.open("c:/Users/Huawei/Desktop/file4.png")
                img.paste(img4, (counter * 100 - 100, 0))
                if counter > 1:
                    img3 = Image.open("c:/Users/Huawei/Desktop/file3.png")
                    img.paste(img3, (counter * 100 - 200, 0))
                if counter > 2:
                    img2 = Image.open("c:/Users/Huawei/Desktop/file2.png")
                    img.paste(img2, (counter * 100 - 300, 0))
                if counter > 3:
                    img1 = Image.open("c:/Users/Huawei/Desktop/file1.png")
                    img.paste(img1, (0, 0))
                img = img.save("c:/Users/Huawei/Desktop/image.png")
                upload = VkUpload(vk_session)
                photo = upload.photo_messages(photos="c:/Users/Huawei/Desktop/image.png")[0]
                attachments = ("photo{}_{}".format(photo["owner_id"], photo['id']))
                vk.messages.send(
                    user_id=event.user_id,
                    random_id=get_random_id(),
                    attachment=attachments,
                    message="\n")
                vk.messages.send(
                    user_id=event.user_id,
                    random_id=get_random_id(),
                    message=information)
                break

def get_weather_in_5_days(vk_session, vk, event):
    response = requests.get("http://api.openweathermap.org/data/2.5/forecast?q=moscow&appid=1e7d1c94703c5b863a60ea656e79de92&lang=ru&units=metric")
    info = response.json()
    date1 = datetime.today() + timedelta(1)
    date1 = normalize_date(date1.day) + "." + normalize_date(date1.month)
    date2 = datetime.today() + timedelta(5)
    date2 = normalize_date(date2.day) + "." + normalize_date(date2.month)
    ms = "Погода в Москве c " + date1 + " по " + date2
    vk.messages.send(user_id=event.user_id, random_id=get_random_id(), message=ms)
    day = []
    night = []
    images = []
    counter = 0
    for i in info["list"]:
        if (counter == 5): break
        if '03:00:00' in i["dt_txt"]:
            night.append(str(round(i['main']['temp'])) + ' °C')
        elif '15:00:00' in i["dt_txt"]:
            counter += 1
            day.append(str(round(i['main']['temp'])) + ' °C')
            images.append(i['weather'][0]['icon'])
    counter = 0
    for i in images:
        counter += 1
        with open("file{}.png".format(counter), "wb") as f:
            image=requests.get("http://openweathermap.org/img/wn/{}@2x.png".format(i), stream=True)
            f.write(image.content)
    img = Image.new('RGBA', (500, 100))
    img1 = Image.open("file1.png")
    img2 = Image.open("file2.png")
    img3 = Image.open("file3.png")
    img4 = Image.open("file4.png")
    img5 = Image.open("file4.png")
    img.paste(img1, (0, 0))
    img.paste(img2, (100, 0))
    img.paste(img3, (200, 0))
    img.paste(img4, (300, 0))
    img.paste(img5, (400, 0))
    img = img.save("image.png")
    upload = VkUpload(vk_session)
    photo = upload.photo_messages(photos="image.png")[0]
    attachments = ("photo{}_{}".format(photo["owner_id"], photo['id']))
    vk.messages.send(
        user_id=event.user_id,
        random_id=get_random_id(),
        attachment=attachments,
        message="\n")
    ms = "\n" + 'ДЕНЬ: '
    for i in day:
        ms += i + " // "
    ms += "\n" + "НОЧЬ: "
    for i in night:
        ms += i + " // "
    vk.messages.send(
        user_id=event.user_id,
        random_id=get_random_id(),
        message=ms)

def bofort_scale(speed):
    if speed <= 0.2: return "Штиль"
    if speed <= 1.5: return "Тихий"
    if speed <= 3.3: return "Лёгкий"
    if speed <= 5.4: return "Слабый"
    if speed <= 7.9: return "Умеренный"
    if speed <= 10.7: return "Свежий"
    if speed <= 13.8: return "Сильный"
    if speed <= 17.1: return "Крепкий"
    if speed <= 20.7: return "Очень крепкий"
    if speed <= 24.4: return "Шторм"
    if speed <= 28.4: return "Сильный шторм"
    if speed <= 32.6: return "Жестокий шторм"
    return "Ураган"

def rumb(deg):
    if deg <= 22.5 or deg >= 337.5: return "северный"
    if deg <= 67.5: return "северо-восточный"
    if deg <= 112.5: return "восточный"
    if deg <= 157.5: return "юго-восточный"
    if deg <= 202.5: return "южный"
    if deg <= 247.5: return "юго-западный"
    if deg <= 292.5: return "западный"
    return "северо-западный"

def get_weather_now():
    translator = Translator(to_lang="ru")
    weathet_url = "http://api.openweathermap.org/data/2.5/weather?q=moscow&appid=" + WEATHER_KEY + "&units=metric"
    response = requests.get(weathet_url)
    info = response.json()
    information = ""
    result = translator.translate(info["weather"][0]["main"])
    information += "Погода в Москве: " # + result + "\n"
    result = translator.translate(info["weather"][0]["description"])
    information += result + ', температура: ' + str(round(info["main"]["temp_min"])) + ' - ' + str(round(info["main"]["temp_max"])) + '\n'
    information += "Давление: " + str(info["main"]["pressure"]) + " мм.рт.ст., влажность: " + str(info["main"]["humidity"]) + '\n'
    speed = info["wind"]["speed"]
    deg = info["wind"]["deg"]
    information += "Ветер: " + bofort_scale(speed) + ", " + str(speed) + "м/с, " + rumb(deg)
    return information

def get_coronavirus_stat(vk, vk_session, event):
    response = requests.get("https://coronavirusstat.ru/country/russia/")
    soup = BeautifulSoup(response.text, "html.parser")
    result = soup.findAll('table')[0].find("tbody").findAll("td")
    s = soup.findAll('body')[0].find("h6").find('strong').text + '\n'
    k = 0
    for i in result[0]:
        if k == 0:
            s += "Активных: " + str(i)
        elif k == 1:
            s += "({} за сегодня)".format(i.text) + '\n'
        else:
            break
        k += 1
    k = 0
    for i in result[1]:
        if k == 0:
            s += "Вылечено: " + str(i)

        elif k == 1:
            s += "({} за сегодня)\n".format(i.text)
        else:
            break
        k += 1
    k = 0
    for i in result[2]:
        if k == 0:
            s += "Умерло: " + str(i)
        elif k == 1:
            s += "({} за сегодня)\n".format(i.text)
        else:
            break
        k += 1
    k = 0
    for i in result[3]:
        if k == 0:
            s += "Случаев: " + str(i)

        elif k == 1:
            s += "({} за сегодня)".format(i.text)
        else:
            break
        k += 1
    result = soup.findAll('table')[0].find("tbody").findAll("td", {"class": "d-none d-sm-block"})
    infected = []
    k = 0
    for i in result:
        if k < 10:
            if i.find("span", {"class": "badge badge-danger"}):
                infected.append(int(i.find("span", {"class": "badge badge-danger"}).text))
                k += 1
        else:
            break
    
    result = soup.findAll('table')[0].find("tbody").findAll("span", {"class": "badge badge-success"})

    cured = []
    k = 0
    for i in result:
        if k < 20 and k % 2 == 1:
            print(i.text)
            cured.append(int(i.text))
        elif k > 20:
            break
        k += 1
    result = soup.findAll('table')[0].find("tbody").findAll("th")
    data = []
    k = 0
    for i in result:
        if k < 10:
            print(i.text)
            data.append(i.text[:5])
        else:
            break
        k += 1
    a = np.array([cured, infected])
    barWidth = 0.25
    fig = plt.subplots()

    br1 = np.arange(len(cured))
    br2 = [x + barWidth for x in br1]

    plt.bar(br1, cured, color='b', width=barWidth,
            edgecolor='grey', label='Выздоровевшие')
    plt.bar(br2, infected, color='r', width=barWidth,
            edgecolor='grey', label='Заболевшие')

    plt.xlabel('', fontweight='bold', fontsize=15)
    plt.ylabel('Кол-во', fontweight='bold', fontsize=15)
    plt.xticks([r + barWidth for r in range(len(cured))],
                data)

    plt.legend()
    plt.savefig("grafic.png")

    upload = VkUpload(vk_session)
    photo = upload.photo_messages(photos="grafic.png")[0]
    attachments = []
    attachments.append("photo{}_{}".format(photo["owner_id"], photo['id']))
    vk.messages.send(
        user_id=event.user_id,
        random_id=get_random_id(),
        attachment=attachments[0],
        message=s)

def get_coronavirus_stat_by_region(region, vk, event):
    response = requests.get("https://coronavirusstat.ru")
    soup = BeautifulSoup(response.text, "html.parser")
    result = soup.findAll('div', {'class': "row border border-bottom-0 c_search_row"})
    a = ''
    region = region.capitalize()
    for i in range(len(result)):
        if region in result[i].find('a').text:
            s = result[i].findAll("span", {"class": "dline"})
            a += "Активных: "+s[0].text+"\n"
            a += "Вылечено: "+s[1].text+"\n"
            a += "Умерло: "+s[2].text+"\n"
            s = result[i].findAll("div", {"class": "h6 m-0"})
            a += "Заразилось: "+s[0].text[2:]
            break
    if a == "":
        vk.messages.send(
        user_id=event.user_id,
        random_id=get_random_id(),
        message="Регион не найден")
    else:
        vk.messages.send(
        user_id=event.user_id,
        random_id=get_random_id(),
        message=a[:len(a) - 11])

def main():
    global SHEET
    global NUM_COLS
    global NUM_ROWS
    f = open("c:/Users/Huawei/Desktop/1c.xlsx", "wb")
    resp = requests.get(get_links()[0])
    f.write(resp.content)
    f.close()
    f = open("c:/Users/Huawei/Desktop/2c.xlsx", "wb")
    resp = requests.get(get_links()[1])
    f.write(resp.content)
    f.close()
    f = open("c:/Users/Huawei/Desktop/3c.xlsx", "wb")
    resp = requests.get(get_links()[2])
    f.write(resp.content)
    f.close()
    book1 = openpyxl.load_workbook("c:/Users/Huawei/Desktop/1c.xlsx") # открытие файла
    book2 = openpyxl.load_workbook("c:/Users/Huawei/Desktop/2c.xlsx") # открытие файла
    book3 = openpyxl.load_workbook("c:/Users/Huawei/Desktop/3c.xlsx") # открытие файла

    
    SHEET = [book1.active, book2.active, book3.active] # активный лист
    NUM_COLS = [SHEET[i].max_column for i in range(3)] # количество столбцов
    NUM_ROWS = [SHEET[i].max_row for i in range(3)] # количество строк
    vk_session = vk_api.VkApi(token='0021cf58bb0d25bd735c01918b092284fa1f13be6e3943558152c8c4a1299c9c7895bb569a308481ee3c8')
    vk = vk_session.get_api()
    longpoll = VkLongPoll(vk_session)
    keyboard = VkKeyboard(one_time=True)
    flag_schedule = False
    flag_weather = False
    flag_teacher = False
    last_message = ""
    teacher_name = ""
    current_group = "ИКБО-08-21"


    for event in longpoll.listen():
        if event.type == VkEventType.MESSAGE_NEW:
            full_message = list(event.text.split())

        if event.type == VkEventType.MESSAGE_NEW and event.text.lower() == "start":
            ms = 'Для начала работы с ботом напишите привет. Функции бота можно посомотреть, введя БОТ' + "\n" + 'Узнать расписание для конкретной группы можно, введя запрос в формате БОТ + "номер группы" или же БОТ "день недели" "номер группы" ' + "\n" + 'При вводе Коронавирус "название области" выведется статистика по коронавирусу в текущем регионе'
            vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
        
        elif event.type == VkEventType.MESSAGE_NEW and event.text.lower() == "привет" and event.to_me:
            first_message(vk, event)

        elif event.type == VkEventType.MESSAGE_NEW and event.text.lower() == "бот":
            keyboard = VkKeyboard(one_time=True)
            call_keyboard_first(keyboard, vk, event)

        elif event.type == VkEventType.MESSAGE_NEW and event.text.lower() == "получить расписание":
            keyboard = VkKeyboard(one_time=True)
            call_keyboard_schedule(keyboard, vk, event)
            flag_schedule = True

        elif event.type == VkEventType.MESSAGE_NEW and event.text.lower() == "получить погоду":
            keyboard = VkKeyboard(one_time=True)
            call_keyboard_weather(keyboard, vk, event)
            flag_weather = True

        elif event.type == VkEventType.MESSAGE_NEW and len(event.text) == 3 and "бот" in event.text.lower() and event.text.lower().split()[1] in DAYS:
            group = event.text.split()[2].upper()
            vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=get_formatted_schedule(group, event.text.lower().split()[1]))

        elif event.type == VkEventType.MESSAGE_NEW and len(event.text) == 2 and "бот" in event.text.lower() and event.text.lower().split()[1] in DAYS:
            vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=get_formatted_schedule(current_group, event.text.lower().split()[1]))


        elif event.type == VkEventType.MESSAGE_NEW and flag_teacher and event.to_me:
            s = get_formatted_teacher(teacher_name, event.text.lower())
            vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=s)
            flag_teacher = False


        elif event.type == VkEventType.MESSAGE_NEW and len(full_message) == 1 and full_message[0][0].lower() == "и":
            current_group = full_message[0].upper()
            ms = "Я запомнил, что ты из группы " + full_message[0].upper()
            vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)

        elif event.type == VkEventType.MESSAGE_NEW and flag_schedule and event.to_me:
            if event.text != "какая неделя?" and event.text != 'какая группа?':
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=get_formatted_schedule(current_group, event.text))
            elif event.text == "какая неделя?":
                first_day = datetime(2022, 2, 7)
                current_week = datetime.today() - first_day
                current_week = 1 + int(current_week.days) // 7
                ms = "Сейчас идёт " + str(current_week) + " неделя"
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
            else:
                ms = "Показываю расписание группы " + current_group
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
            flag_schedule = False

        elif event.type == VkEventType.MESSAGE_NEW and len(full_message) == 2 and full_message[0].lower() == "бот":
            if full_message[1][0].lower() == "и":
                current_group = full_message[1].upper()
                ms = "Выбрана группа " + current_group
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
                ms = "Получить расписание"
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
            else:
                print(full_message)
                ms = get_formatted_schedule(current_group, full_message[1].lower())
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)

        elif event.type == VkEventType.MESSAGE_NEW and len(full_message) == 3 and full_message[0].lower() == "бот":
            ms = get_formatted_schedule(full_message[2].upper(), full_message[1].lower())
            vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
        
        elif event.type == VkEventType.MESSAGE_NEW and event.text.lower() == "получить статистику по коронавирусу":
            get_coronavirus_stat(vk, vk_session, event)

        elif event.type == VkEventType.MESSAGE_NEW and len(full_message) == 2 and full_message[0].lower() == "коронавирус":
            get_coronavirus_stat_by_region(full_message[1].lower(), vk, event)

        elif event.type == VkEventType.MESSAGE_NEW and flag_weather and event.text.lower() == "сейчас":
            ms = get_weather_now()
            vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
            flag_schedule = False

        elif event.type == VkEventType.MESSAGE_NEW and flag_weather and (event.text.lower() == "сегодня" or event.text.lower() == "завтра"):
            get_weather_today(event.text.lower(), vk_session, vk, event)
            flag_schedule = False

        elif event.type == VkEventType.MESSAGE_NEW and flag_weather and (event.text.lower() == "на 5 дней"):
            get_weather_in_5_days(vk_session, vk, event)
            flag_schedule = False
        
        elif event.type == VkEventType.MESSAGE_NEW and len(full_message) == 2 and full_message[0].lower() == "найти":
            s = get_teacher_schedule(event.text.lower().split()[1])
            keyboard = VkKeyboard(one_time=True)
            vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message="Показать расписание преподавателя " + s[1])
            call_keyboard_teacher(keyboard, vk, event)
            flag_teacher = True
            teacher_name = s[1]

        elif event.type == VkEventType.MESSAGE_NEW and event.to_me:
            ms = "Неизвестная команда. Для списка команд напишите БОТ"
            vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)

if __name__ == '__main__':
    main()
